import sys
import os
import json
import logging
import pandas as pd
import MetaTrader5 as mt5
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QLabel, QLineEdit, QComboBox, QPushButton, QRadioButton, 
                             QDateEdit, QSpinBox, QMessageBox, QFileDialog, 
                             QProgressBar, QListWidget, QListWidgetItem, QFrame, 
                             QSplitter, QToolButton, QInputDialog, QDialog, 
                             QDialogButtonBox, QListWidget, QGridLayout, QMenuBar, 
                             QMenu, QAction, QFontDialog, QColorDialog, QCheckBox)
from PyQt5.QtCore import QDate, Qt, QThread, pyqtSignal, QLocale, QUrl
from PyQt5.QtGui import QFont, QPalette, QColor, QIcon, QDesktopServices
import pytz
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
import matplotlib.dates as mdates
from mplfinance.original_flavor import candlestick_ohlc

# Configure logging
logging.basicConfig(
    filename='mt5_downloader.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

class Translator:
    """Handles language translations for the application"""
    def __init__(self):
        self.translations = {
            'en': {
                'app_title': "MT5 Historical Data Downloader",
                'symbols_label': "Symbols (comma-separated):",
                'symbol_management': "Symbol Management",
                'load_preset': "Load symbol preset...",
                'save_preset': "Save Preset",
                'load_watchlist': "Load Watchlist",
                'settings': "Settings",
                'timeframes': "Timeframes:",
                'export_format': "Export Format:",
                'date_range': "Date Range",
                'specific_range': "Specific Date Range",
                'days_back': "Days Back from Today",
                'start_date': "Start Date:",
                'end_date': "End Date:",
                'days_label': "Days back:",
                'output': "Output",
                'output_path': "Output Path:",
                'columns_label': "Columns to include in export:",
                'plot_label': "Select Symbol to Plot:",
                'plot_btn': "Plot Selected Symbol",
                'download_btn': "Download Data",
                'stop_btn': "Stop Download",
                'browse_btn': "Browse",
                'theme_tooltip': "Toggle Theme",
                'no_data': "No data to display",
                'price_chart': "Price Chart",
                'invalid_symbol': "Invalid symbol detected between commas",
                'no_symbols': "Please enter at least one valid symbol",
                'no_timeframes': "Please select at least one timeframe",
                'invalid_range': "Start date must be before end date",
                'no_columns': "Please select at least one column to export",
                'download_starting': "Starting download...",
                'download_complete': "Download completed successfully",
                'no_data_received': "Download complete but no data received",
                'error_occurred': "Error occurred",
                'mt5_connect_error': "Failed to connect to MT5. Please ensure MT5 is running.",
                'no_watchlist': "No symbols found in Market Watch",
                'preset_name': "Enter preset name:",
                'preset_saved': "Preset '{}' saved",
                'preset_loaded': "Loaded preset '{}'",
                'watchlist_loaded': "Loaded {} symbols from watchlist",
                'confirm_exit': "A download is currently running. Are you sure you want to quit?",
                'theme_message': "Dark theme is currently the only available option",
                'select_symbols': "Select symbols to import:",
                'watchlist_title': "Select Watchlist Symbols",
                'tooltip_symbols': "Enter symbols separated by commas (e.g., XAUUSD,EURUSD)",
                'tooltip_timeframes': "Select one or more timeframes for data download",
                'tooltip_export_format': "Choose the file format for exported data",
                'tooltip_date_range': "Select a specific date range or days back",
                'tooltip_output_path': "Specify the file path for exported data",
                'tooltip_columns': "Select columns to include in the exported file",
                'tooltip_plot': "Select a symbol to plot its price chart",
                'tooltip_download': "Start downloading data for selected symbols",
                'tooltip_stop': "Stop the current download process",
                'tooltip_browse': "Browse to select the output file location",
                'tooltip_save_preset': "Save the current symbols as a preset",
                'tooltip_load_watchlist': "Load symbols from MT5 Market Watch",
                'menu_file': "File",
                'menu_settings': "Settings",
                'menu_help': "Help",
                'action_save_preset': "Save Preset",
                'action_load_watchlist': "Load Watchlist",
                'action_exit': "Exit",
                'action_select_font': "Select Font...",
                'action_font_size': "Font Size",
                'action_toggle_language': "Toggle Language",
                'action_chart_settings': "Chart Settings...",
                'action_about': "About",
                'action_view_log': "View Log",
                'about_title': "About MT5 Historical Data Downloader",
                'about_text': "Version 1.0\nA tool for downloading historical data from MetaTrader 5.\nDeveloped by xAI.",
                'chart_settings_title': "Chart Settings",
                'up_color': "Up Candle Color:",
                'down_color': "Down Candle Color:",
                'show_grid': "Show Grid",
                'show_volume': "Show Volume Overlay",
            },
            'fa': {
                'app_title': "دانلودگر داده‌های تاریخی متاتریدر ۵",
                'symbols_label': "نمادها (جداشده با کاما):",
                'symbol_management': "مدیریت نمادها",
                'load_preset': "بارگیری پیش‌تنظیم نمادها...",
                'save_preset': "ذخیره پیش‌تنظیم",
                'load_watchlist': "بارگیری واچ‌لیست",
                'settings': "تنظیمات",
                'timeframes': "تایم‌فریم‌ها:",
                'export_format': "فرمت خروجی:",
                'date_range': "بازه زمانی",
                'specific_range': "بازه تاریخ مشخص",
                'days_back': "روز قبل از امروز",
                'start_date': "تاریخ شروع:",
                'end_date': "تاریخ پایان:",
                'days_label': "تعداد روز:",
                'output': "خروجی",
                'output_path': "مسیر ذخیره:",
                'columns_label': "ستون‌های مورد نظر برای export:",
                'plot_label': "انتخاب نماد برای نمایش:",
                'plot_btn': "نمایش نماد انتخاب شده",
                'download_btn': "دانلود داده‌ها",
                'stop_btn': "توقف دانلود",
                'browse_btn': "مرور",
                'theme_tooltip': "تغییر تم",
                'no_data': "داده‌ای برای نمایش وجود ندارد",
                'price_chart': "نمودار قیمت",
                'invalid_symbol': "نماد خالی بین کاماها وجود دارد",
                'no_symbols': "لطفا حداقل یک نماد معتبر وارد کنید",
                'no_timeframes': "لطفا حداقل یک تایم‌فریم انتخاب کنید",
                'invalid_range': "تاریخ شروع باید قبل از تاریخ پایان باشد",
                'no_columns': "لطفا حداقل یک ستون برای export انتخاب کنید",
                'download_starting': "در حال شروع دانلود...",
                'download_complete': "دانلود با موفقیت انجام شد",
                'no_data_received': "دانلود کامل شد اما داده‌ای دریافت نشد",
                'error_occurred': "خطا رخ داده است",
                'mt5_connect_error': "اتصال به متاتریدر ۵ ناموفق بود. لطفا از اجرا بودن متاتریدر اطمینان حاصل کنید.",
                'no_watchlist': "هیچ نمادی در واچ‌لیست یافت نشد",
                'preset_name': "نام پیش‌تنظیم را وارد کنید:",
                'preset_saved': "پیش‌تنظیم '{}' ذخیره شد",
                'preset_loaded': "پیش‌تنظیم '{}' بارگیری شد",
                'watchlist_loaded': "{} نماد از واچ‌لیست بارگیری شد",
                'confirm_exit': "دانلود در حال انجام است. آیا مطمئنید که می‌خواهید خارج شوید؟",
                'theme_message': "در حال حاضر فقط تم تیره موجود است",
                'select_symbols': "نمادهای مورد نظر را انتخاب کنید:",
                'watchlist_title': "انتخاب نمادهای واچ‌لیست",
                'tooltip_symbols': "نمادها را با کاما جدا کنید (مثال: XAUUSD,EURUSD)",
                'tooltip_timeframes': "یک یا چند تایم‌فریم برای دانلود انتخاب کنید",
                'tooltip_export_format': "فرمت فایل برای داده‌های خروجی را انتخاب کنید",
                'tooltip_date_range': "بازه تاریخ مشخص یا تعداد روز قبل را انتخاب کنید",
                'tooltip_output_path': "مسیر فایل خروجی را مشخص کنید",
                'tooltip_columns': "ستون‌های مورد نظر برای فایل خروجی را انتخاب کنید",
                'tooltip_plot': "نمادی را برای نمایش نمودار قیمت انتخاب کنید",
                'tooltip_download': "شروع دانلود داده‌ها برای نمادهای انتخاب شده",
                'tooltip_stop': "توقف فرآیند دانلود جاری",
                'tooltip_browse': "انتخاب مکان فایل خروجی",
                'tooltip_save_preset': "ذخیره نمادهای فعلی به عنوان پیش‌تنظیم",
                'tooltip_load_watchlist': "بارگیری نمادها از واچ‌لیست متاتریدر ۵",
                'menu_file': "فایل",
                'menu_settings': "تنظیمات",
                'menu_help': "کمک",
                'action_save_preset': "ذخیره پیش‌تنظیم",
                'action_load_watchlist': "بارگیری واچ‌لیست",
                'action_exit': "خروج",
                'action_select_font': "انتخاب فونت...",
                'action_font_size': "اندازه فونت",
                'action_toggle_language': "تغییر زبان",
                'action_chart_settings': "تنظیمات نمودار...",
                'action_about': "درباره",
                'action_view_log': "مشاهده لاگ",
                'about_title': "درباره دانلودگر داده‌های تاریخی متاتریدر ۵",
                'about_text': "نسخه ۱.۰\nابزاری برای دانلود داده‌های تاریخی از متاتریدر ۵.\nتوسعه‌یافته توسط xAI.",
                'chart_settings_title': "تنظیمات نمودار",
                'up_color': "رنگ کندل صعودی:",
                'down_color': "رنگ کندل نزولی:",
                'show_grid': "نمایش شبکه",
                'show_volume': "نمایش حجم معاملات",
            }
        }
        self.current_lang = 'en'
        
    def set_language(self, lang):
        """Set the current language (en/fa)"""
        if lang in self.translations:
            self.current_lang = lang
            return True
        return False
    
    def tr(self, text_key):
        """Translate text based on current language"""
        return self.translations[self.current_lang].get(text_key, text_key)

class ChartSettingsDialog(QDialog):
    """Dialog for customizing chart settings"""
    def __init__(self, parent):
        super().__init__(parent)
        self.translator = parent.translator
        self.setWindowTitle(self.translator.tr('chart_settings_title'))
        self.setMinimumWidth(300)
        layout = QVBoxLayout()
        
        # Up candle color
        up_layout = QHBoxLayout()
        self.up_color_label = QLabel(self.translator.tr('up_color'))
        up_layout.addWidget(self.up_color_label)
        self.up_color_btn = QPushButton()
        self.up_color_btn.setFixedSize(40, 40)
        self.up_color_btn.clicked.connect(self.select_up_color)
        up_layout.addWidget(self.up_color_btn)
        layout.addLayout(up_layout)
        
        # Down candle color
        down_layout = QHBoxLayout()
        self.down_color_label = QLabel(self.translator.tr('down_color'))
        down_layout.addWidget(self.down_color_label)
        self.down_color_btn = QPushButton()
        self.down_color_btn.setFixedSize(40, 40)
        self.down_color_btn.clicked.connect(self.select_down_color)
        down_layout.addWidget(self.down_color_btn)
        layout.addLayout(down_layout)
        
        # Grid toggle
        self.grid_check = QCheckBox(self.translator.tr('show_grid'))
        layout.addWidget(self.grid_check)
        
        # Volume toggle
        self.volume_check = QCheckBox(self.translator.tr('show_volume'))
        layout.addWidget(self.volume_check)
        
        # Buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        
        self.setLayout(layout)
        
        # Load current settings
        self.up_color = QColor(parent.candle_colors['up'])
        self.down_color = QColor(parent.candle_colors['down'])
        self.grid_check.setChecked(parent.show_grid)
        self.volume_check.setChecked(parent.show_volume)
        self.update_color_buttons()
        
    def select_up_color(self):
        color = QColorDialog.getColor(self.up_color, self, "Select Up Candle Color")
        if color.isValid():
            self.up_color = color
            self.update_color_buttons()
            
    def select_down_color(self):
        color = QColorDialog.getColor(self.down_color, self, "Select Down Candle Color")
        if color.isValid():
            self.down_color = color
            self.update_color_buttons()
            
    def update_color_buttons(self):
        self.up_color_btn.setStyleSheet(f"background-color: {self.up_color.name()}; border: 1px solid #616161;")
        self.down_color_btn.setStyleSheet(f"background-color: {self.down_color.name()}; border: 1px solid #616161;")

class CandlestickChart(FigureCanvas):
    def __init__(self, parent=None, width=5, height=4, dpi=100):
        self.fig = Figure(figsize=(width, height), dpi=dpi, facecolor='#212121')
        self.ax = self.fig.add_subplot(111)
        super().__init__(self.fig)
        self.setParent(parent)
        self.dark_mode = True
        self.translator = Translator()
        
    def plot_candles(self, data, symbol="", candle_colors=None, show_grid=True, show_volume=False):
        """Plot candlestick chart from OHLC data with enhanced styling"""
        self.fig.clear()
        
        if show_volume and 'Volume' in data.columns:
            self.ax = self.fig.add_subplot(211)
            self.ax_volume = self.fig.add_subplot(212, sharex=self.ax)
        else:
            self.ax = self.fig.add_subplot(111)
            self.ax_volume = None
            
        if data.empty:
            self.ax.text(0.5, 0.5, self.translator.tr('no_data'), 
                        ha='center', va='center', fontsize=12,
                        color='#FFFFFF')
            self.draw()
            return
        
        # Convert dates to matplotlib format
        dates = mdates.date2num(data['Date'])
        
        # Prepare OHLC data
        ohlc = list(zip(dates, data['Open'], data['High'], data['Low'], data['Close']))
        
        # Plot candlesticks with user-defined colors
        candle_colors = candle_colors or {'up': '#4CAF50', 'down': '#F44336'}
        candlestick_ohlc(self.ax, ohlc, width=0.6, 
                        colorup=candle_colors['up'], 
                        colordown=candle_colors['down'], 
                        alpha=0.9)
        
        # Plot volume if enabled
        if self.ax_volume and 'Volume' in data.columns:
            self.ax_volume.bar(dates, data['Volume'], color='#6200EA', alpha=0.6)
            self.ax_volume.set_ylabel('Volume', color='#FFFFFF', fontsize=10)
            self.ax_volume.tick_params(colors='#FFFFFF', labelsize=8)
            self.ax_volume.set_facecolor('#212121')
            for spine in self.ax_volume.spines.values():
                spine.set_color('#FFFFFF')
            self.fig.subplots_adjust(hspace=0)
        
        # Format x-axis with better date display
        self.ax.xaxis_date()
        self.fig.autofmt_xdate()
        locator = mdates.AutoDateLocator()
        formatter = mdates.ConciseDateFormatter(locator)
        self.ax.xaxis.set_major_locator(locator)
        self.ax.xaxis.set_major_formatter(formatter)
        
        # Flat chart styling
        bg_color = '#212121'
        text_color = '#FFFFFF'
        grid_color = '#424242'
        
        self.ax.grid(show_grid, linestyle='--', alpha=0.7, color=grid_color)
        self.ax.set_facecolor(bg_color)
        self.fig.patch.set_facecolor(bg_color)
        self.ax.tick_params(colors=text_color, labelsize=10)
        self.ax.xaxis.label.set_color(text_color)
        self.ax.yaxis.label.set_color(text_color)
        self.ax.title.set_color(text_color)
        
        for spine in self.ax.spines.values():
            spine.set_color(text_color)
        
        # Set title with symbol
        title = f"{symbol} {self.translator.tr('price_chart')}" if symbol else self.translator.tr('price_chart')
        self.ax.set_title(title, pad=20, fontsize=14, fontfamily='Roboto')
        self.ax.set_ylabel('Price', labelpad=10, fontsize=12, fontfamily='Roboto')
        
        self.draw()

class DataDownloadThread(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(dict)  # Emits dict of dataframes with full OHLC data
    error = pyqtSignal(str)
    log_message = pyqtSignal(str, str)  # message, level

    def __init__(self, symbols, timeframes, start_dt, end_dt, output_file, export_format, selected_columns):
        super().__init__()
        self.symbols = symbols
        self.timeframes = timeframes
        self.start_dt = start_dt
        self.end_dt = end_dt
        self.output_file = output_file
        self.export_format = export_format
        self.selected_columns = selected_columns
        self._is_running = True

    def run(self):
        try:
            self.log_message.emit("Attempting to initialize MT5 connection...", "INFO")
            if not mt5.initialize():
                error_msg = f"Failed to initialize MT5 connection. Error: {mt5.last_error()}"
                self.log_message.emit(error_msg, "ERROR")
                self.error.emit(error_msg)
                return

            try:
                self.log_message.emit("Getting available symbols...", "INFO")
                available_symbols = [s.name for s in mt5.symbols_get()]
                self.log_message.emit(f"Found {len(available_symbols)} available symbols", "INFO")
                
                result_data = {}  # Stores full OHLC data for charting
                export_data = {}  # Stores data with user-selected columns for export
                
                tf_mapping = {
                    'M1': mt5.TIMEFRAME_M1, 'M5': mt5.TIMEFRAME_M5, 'M15': mt5.TIMEFRAME_M15,
                    'M30': mt5.TIMEFRAME_M30, 'H1': mt5.TIMEFRAME_H1, 'H4': mt5.TIMEFRAME_H4,
                    'D1': mt5.TIMEFRAME_D1, 'W1': mt5.TIMEFRAME_W1, 'MN1': mt5.TIMEFRAME_MN1,
                }

                total_tasks = len(self.symbols) * len(self.timeframes)
                completed_tasks = 0
                self.log_message.emit(f"Starting download of {total_tasks} symbol/timeframe combinations", "INFO")
                
                for symbol in self.symbols:
                    if not self._is_running:
                        break
                        
                    exact_symbol = next((s for s in available_symbols if s.lower() == symbol.lower()), None)
                    
                    if not exact_symbol:
                        msg = f"Symbol {symbol} not found (case mismatch)"
                        self.log_message.emit(msg, "WARNING")
                        completed_tasks += len(self.timeframes)
                        self.progress.emit(int((completed_tasks / total_tasks) * 100))
                        continue

                    for timeframe in self.timeframes:
                        if not self._is_running:
                            break
                            
                        try:
                            self.log_message.emit(f"Downloading {exact_symbol} {timeframe} data...", "INFO")
                            rates = mt5.copy_rates_range(
                                exact_symbol, 
                                tf_mapping[timeframe], 
                                self.start_dt, 
                                self.end_dt
                            )

                            if rates is None or len(rates) == 0:
                                error_msg = f"No data returned for {exact_symbol} {timeframe}: {mt5.last_error()}"
                                self.log_message.emit(error_msg, "WARNING")
                                completed_tasks += 1
                                self.progress.emit(int((completed_tasks / total_tasks) * 100))
                                continue

                            # Create full OHLC dataframe for charting
                            full_df = pd.DataFrame(rates)
                            if full_df.empty:
                                msg = f"Empty dataframe for symbol {exact_symbol} {timeframe}"
                                self.log_message.emit(msg, "WARNING")
                                completed_tasks += 1
                                self.progress.emit(int((completed_tasks / total_tasks) * 100))
                                continue

                            # Convert and rename columns
                            full_df['time'] = pd.to_datetime(full_df['time'], unit='s')
                            if full_df['time'].dt.tz is None:
                                full_df['time'] = full_df['time'].dt.tz_localize('UTC')
                                
                            full_df = full_df.rename(columns={
                                'time': 'Date', 'open': 'Open', 'high': 'High',
                                'low': 'Low', 'close': 'Close', 'tick_volume': 'Volume',
                                'spread': 'Spread', 'real_volume': 'RealVolume'
                            })

                            # Store full OHLC data for charting with timeframe in key
                            result_data[f"{exact_symbol}_{timeframe}"] = full_df[['Date', 'Open', 'High', 'Low', 'Close', 'Volume']]
                            
                            # Create export dataframe with user-selected columns
                            export_df = full_df.copy()
                            available_cols = [col for col in self.selected_columns if col in export_df.columns]
                            if not available_cols:
                                msg = f"No valid columns selected for {exact_symbol} {timeframe}"
                                self.log_message.emit(msg, "WARNING")
                                completed_tasks += 1
                                self.progress.emit(int((completed_tasks / total_tasks) * 100))
                                continue
                            
                            export_df = export_df[available_cols]
                            
                            # For Excel export, we'll handle all data together after the loop
                            if self.export_format == "xlsx":
                                export_data[f"{exact_symbol}_{timeframe}"] = export_df
                            elif self.output_file:
                                # Single symbol/timeframe or CSV export
                                if len(self.symbols) == 1 and len(self.timeframes) == 1 and not os.path.isdir(self.output_file):
                                    filename = self.output_file
                                else:
                                    base_dir = os.path.dirname(self.output_file) if self.output_file else "."
                                    if not os.path.exists(base_dir):
                                        os.makedirs(base_dir)
                                    filename = os.path.join(
                                        base_dir,
                                        f"{exact_symbol}_{timeframe}_"
                                        f"{self.start_dt.strftime('%Y%m%d')}_to_"
                                        f"{self.end_dt.strftime('%Y%m%d')}.{self.export_format}"
                                    )
                                
                                try:
                                    if self.export_format == "xlsx":
                                        if 'Date' in export_df.columns:
                                            export_df['Date'] = export_df['Date'].dt.tz_localize(None)
                                        export_df.to_excel(filename, index=False, engine='openpyxl')
                                    else:
                                        export_df.to_csv(filename, index=False)
                                    self.log_message.emit(f"Successfully saved {len(export_df)} rows to {filename}", "INFO")
                                except Exception as e:
                                    error_msg = f"Failed to save file: {e}"
                                    self.log_message.emit(error_msg, "ERROR")
                                    completed_tasks += 1
                                    self.progress.emit(int((completed_tasks / total_tasks) * 100))
                                    continue

                            completed_tasks += 1
                            progress = int((completed_tasks / total_tasks) * 100)
                            self.progress.emit(progress)
                            self.log_message.emit(f"Progress: {progress}%", "INFO")
                            self.msleep(100)  # Small delay to allow UI updates

                        except Exception as e:
                            error_msg = f"Error processing {exact_symbol} {timeframe}: {str(e)}"
                            self.log_message.emit(error_msg, "ERROR")
                            completed_tasks += 1
                            self.progress.emit(int((completed_tasks / total_tasks) * 100))
                            continue

                # Handle multi-symbol/multi-timeframe Excel export
                if self._is_running and self.export_format == "xlsx" and len(export_data) > 0 and self.output_file:
                    try:
                        self.log_message.emit("Preparing Excel workbook...", "INFO")
                        wb = Workbook()
                        if len(wb.sheetnames) > 0:
                            wb.remove(wb[wb.sheetnames[0]])
                        
                        for sheet_name, df in export_data.items():
                            if 'Date' in df.columns:
                                df['Date'] = df['Date'].dt.tz_localize(None)
                            
                            ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
                            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                                for c_idx, value in enumerate(row):
                                    ws.cell(row=r_idx+1, column=c_idx+1, value=value)
                        
                        wb.save(self.output_file)
                        self.log_message.emit(f"Successfully saved {len(export_data)} sheets to {self.output_file}", "INFO")
                    except Exception as e:
                        error_msg = f"Failed to save Excel file: {str(e)}"
                        self.log_message.emit(error_msg, "ERROR")
                        self.error.emit(error_msg)

                if self._is_running:
                    self.log_message.emit("Download completed successfully", "INFO")
                    self.finished.emit(result_data)
                
            except Exception as e:
                error_msg = f"Error during download: {str(e)}"
                self.log_message.emit(error_msg, "ERROR")
                self.error.emit(error_msg)
                
            finally:
                self.log_message.emit("Shutting down MT5 connection", "INFO")
                mt5.shutdown()
                
        except Exception as e:
            error_msg = f"Unexpected error: {str(e)}"
            self.log_message.emit(error_msg, "ERROR")
            if mt5.initialize():
                mt5.shutdown()
            self.error.emit(error_msg)

    def stop(self):
        self._is_running = False

class MT5DataDownloader(QMainWindow):
    def __init__(self):
        super().__init__()
        self.translator = Translator()
        self.setWindowTitle(self.translator.tr('app_title'))
        self.setMinimumSize(1200, 700)
        self.dark_mode = True
        self.chart_data = {}
        self.current_chart_symbol = None
        self.download_thread = None
        self.current_font = QFont("Roboto", 12)
        self.candle_colors = {'up': '#4CAF50', 'down': '#F44336'}
        self.show_grid = True
        self.show_volume = False
        self.load_settings()
        self.setup_ui()
        self.apply_dark_theme()
        self.chart.dark_mode = self.dark_mode
        self.chart.translator = self.translator

    def load_settings(self):
        """Load font, font size, and chart settings from settings.json"""
        try:
            if os.path.exists('settings.json'):
                with open('settings.json', 'r') as f:
                    settings = json.load(f)
                    font_name = settings.get('font_name', 'Roboto')
                    font_size = settings.get('font_size', 12)
                    self.current_font = QFont(font_name, font_size)
                    self.candle_colors['up'] = settings.get('up_color', '#4CAF50')
                    self.candle_colors['down'] = settings.get('down_color', '#F44336')
                    self.show_grid = settings.get('show_grid', True)
                    self.show_volume = settings.get('show_volume', False)
                    logging.info(f"Loaded font: {font_name}, size: {font_size}, "
                               f"chart settings: {self.candle_colors}, grid: {self.show_grid}, volume: {self.show_volume}")
        except Exception as e:
            logging.error(f"Error loading settings: {str(e)}")
            self.current_font = QFont("Roboto", 12)  # Fallback

    def save_settings(self):
        """Save font, font size, and chart settings to settings.json"""
        try:
            settings = {
                'font_name': self.current_font.family(),
                'font_size': self.current_font.pointSize(),
                'up_color': self.candle_colors['up'],
                'down_color': self.candle_colors['down'],
                'show_grid': self.show_grid,
                'show_volume': self.show_volume
            }
            with open('settings.json', 'w') as f:
                json.dump(settings, f, indent=4)
            logging.info(f"Saved font: {settings['font_name']}, size: {settings['font_size']}, "
                        f"chart settings: {self.candle_colors}, grid: {self.show_grid}, volume: {self.show_volume}")
        except Exception as e:
            logging.error(f"Error saving settings: {str(e)}")
            QMessageBox.warning(self, "Error", f"Failed to save settings: {str(e)}")

    def update_font(self, font=None, size=None):
        """Update the application font and/or size"""
        try:
            if font:
                self.current_font = font
            if size is not None:
                self.current_font.setPointSize(size)
            
            # Update main window
            self.setFont(self.current_font)
            
            # Update specific widgets with preserved weights
            header_font = QFont(self.current_font)
            header_font.setBold(True)
            header_font.setPointSize(self.current_font.pointSize() + 4)
            self.header_label.setFont(header_font)
            
            status_font = QFont(self.current_font)
            status_font.setPointSize(self.current_font.pointSize() - 2)
            self.status_label.setFont(status_font)
            
            # Update other widgets
            for widget in self.findChildren((QLabel, QPushButton, QComboBox, QLineEdit, QSpinBox, QDateEdit, QRadioButton, QListWidget)):
                widget.setFont(self.current_font)
            
            # Update menu bar
            self.menuBar().setFont(self.current_font)
            
            self.apply_dark_theme()  # Reapply styles to ensure consistency
            self.save_settings()
            logging.info(f"Updated font to {self.current_font.family()}, size {self.current_font.pointSize()}")
        except Exception as e:
            logging.error(f"Error updating font: {str(e)}")
            QMessageBox.warning(self, "Error", f"Failed to update font: {str(e)}")

    def setup_ui(self):
        # Create menu bar
        self.setup_menu_bar()
        
        # Central widget
        central_widget = QWidget()
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Main splitter (horizontal)
        main_splitter = QSplitter(Qt.Horizontal)
        
        # Left panel (controls)
        left_panel = QWidget()
        left_layout = QVBoxLayout()
        left_layout.setContentsMargins(20, 20, 20, 20)
        left_layout.setSpacing(20)
        
        # Header with theme and language toggle
        header_layout = QHBoxLayout()
        
        self.header_label = QLabel(self.translator.tr('app_title'))
        header_font = QFont(self.current_font.family(), self.current_font.pointSize() + 4, QFont.Bold)
        self.header_label.setFont(header_font)
        header_layout.addWidget(self.header_label)
        
        # Language toggle button
        self.lang_btn = QPushButton("EN")
        self.lang_btn.setIcon(QIcon.fromTheme('preferences-desktop-locale'))
        self.lang_btn.setFixedHeight(40)
        self.lang_btn.clicked.connect(self.toggle_language)
        header_layout.addWidget(self.lang_btn)
        
        # Theme toggle button
        self.theme_btn = QToolButton()
        self.theme_btn.setCheckable(True)
        self.theme_btn.setChecked(True)
        self.theme_btn.setIcon(QIcon.fromTheme('color-management'))
        self.theme_btn.setToolTip(self.translator.tr('theme_tooltip'))
        self.theme_btn.setFixedSize(40, 40)
        header_layout.addStretch()
        header_layout.addWidget(self.theme_btn)
        
        left_layout.addLayout(header_layout)
        
        # Main controls grid
        controls_grid = QGridLayout()
        controls_grid.setSpacing(15)
        row = 0
        
        # Symbol input
        self.symbols_label = QLabel(self.translator.tr('symbols_label'))
        self.symbols_label.setFont(self.current_font)
        controls_grid.addWidget(self.symbols_label, row, 0)
        self.symbol_input = QLineEdit("XAUUSD,EURUSD,GBPJPY,BTCUSD")
        self.symbol_input.setToolTip(self.translator.tr('tooltip_symbols'))
        self.symbol_input.setFont(self.current_font)
        controls_grid.addWidget(self.symbol_input, row, 1)
        row += 1
        
        # Symbol management buttons
        symbol_btn_layout = QHBoxLayout()
        self.preset_combo = QComboBox()
        self.preset_combo.setPlaceholderText(self.translator.tr('load_preset'))
        self.preset_combo.setToolTip(self.translator.tr('load_preset'))
        self.preset_combo.setFont(self.current_font)
        self.load_presets()
        self.preset_combo.currentIndexChanged.connect(self.on_preset_selected)
        symbol_btn_layout.addWidget(self.preset_combo)
        
        self.save_preset_btn = QPushButton(self.translator.tr('save_preset'))
        self.save_preset_btn.setIcon(QIcon.fromTheme('document-save'))
        self.save_preset_btn.setToolTip(self.translator.tr('tooltip_save_preset'))
        self.save_preset_btn.setFont(self.current_font)
        self.save_preset_btn.clicked.connect(self.save_symbol_preset)
        symbol_btn_layout.addWidget(self.save_preset_btn)
        
        self.watchlist_btn = QPushButton(self.translator.tr('load_watchlist'))
        self.watchlist_btn.setIcon(QIcon.fromTheme('view-list'))
        self.watchlist_btn.setToolTip(self.translator.tr('tooltip_load_watchlist'))
        self.watchlist_btn.setFont(self.current_font)
        self.watchlist_btn.clicked.connect(self.load_watchlist_symbols)
        symbol_btn_layout.addWidget(self.watchlist_btn)
        
        controls_grid.addLayout(symbol_btn_layout, row, 0, 1, 2)
        row += 1
        
        # Timeframes
        self.timeframes_label = QLabel(self.translator.tr('timeframes'))
        self.timeframes_label.setFont(self.current_font)
        controls_grid.addWidget(self.timeframes_label, row, 0)
        self.tf_list = QListWidget()
        self.tf_list.setSelectionMode(QListWidget.MultiSelection)
        self.tf_list.setToolTip(self.translator.tr('tooltip_timeframes'))
        self.tf_list.setFont(self.current_font)
        for tf in ["M1", "M5", "M15", "M30", "H1", "H4", "D1", "W1", "MN1"]:
            item = QListWidgetItem(tf)
            self.tf_list.addItem(item)
        for i in range(self.tf_list.count()):
            if self.tf_list.item(i).text() == "H1":
                self.tf_list.item(i).setSelected(True)
                break
        self.tf_list.setMaximumHeight(150)
        controls_grid.addWidget(self.tf_list, row, 1)
        row += 1
        
        # Export format
        self.export_format_label = QLabel(self.translator.tr('export_format'))
        self.export_format_label.setFont(self.current_font)
        controls_grid.addWidget(self.export_format_label, row, 0)
        self.format_combo = QComboBox()
        self.format_combo.setToolTip(self.translator.tr('tooltip_export_format'))
        self.format_combo.setFont(self.current_font)
        self.format_combo.addItems(["xlsx", "csv"])
        controls_grid.addWidget(self.format_combo, row, 1)
        row += 1
        
        # Date range
        self.date_range_label = QLabel(self.translator.tr('date_range'))
        self.date_range_label.setFont(self.current_font)
        controls_grid.addWidget(self.date_range_label, row, 0)
        
        date_layout = QVBoxLayout()
        radio_layout = QHBoxLayout()
        self.range_radio = QRadioButton(self.translator.tr('specific_range'))
        self.range_radio.setFont(self.current_font)
        self.days_radio = QRadioButton(self.translator.tr('days_back'))
        self.days_radio.setFont(self.current_font)
        self.days_radio.setChecked(True)
        self.range_radio.toggled.connect(self.toggle_dates)
        radio_layout.addWidget(self.range_radio)
        radio_layout.addWidget(self.days_radio)
        date_layout.addLayout(radio_layout)
        
        self.date_widget = QWidget()
        date_selector_layout = QHBoxLayout()
        self.start_date_label = QLabel(self.translator.tr('start_date'))
        self.start_date_label.setFont(self.current_font)
        date_selector_layout.addWidget(self.start_date_label)
        self.start_date = QDateEdit(QDate.currentDate().addDays(-30))
        self.start_date.setCalendarPopup(True)
        self.start_date.setToolTip(self.translator.tr('tooltip_date_range'))
        self.start_date.setFont(self.current_font)
        date_selector_layout.addWidget(self.start_date)
        self.end_date_label = QLabel(self.translator.tr('end_date'))
        self.end_date_label.setFont(self.current_font)
        date_selector_layout.addWidget(self.end_date_label)
        self.end_date = QDateEdit(QDate.currentDate())
        self.end_date.setCalendarPopup(True)
        self.end_date.setToolTip(self.translator.tr('tooltip_date_range'))
        self.end_date.setFont(self.current_font)
        date_selector_layout.addWidget(self.end_date)
        self.date_widget.setLayout(date_selector_layout)
        date_layout.addWidget(self.date_widget)
        
        self.days_widget = QWidget()
        days_layout = QHBoxLayout()
        self.days_label = QLabel(self.translator.tr('days_label'))
        self.days_label.setFont(self.current_font)
        days_layout.addWidget(self.days_label)
        self.days_spin = QSpinBox()
        self.days_spin.setRange(1, 3650)
        self.days_spin.setValue(30)
        self.days_spin.setToolTip(self.translator.tr('tooltip_date_range'))
        self.days_spin.setFont(self.current_font)
        days_layout.addWidget(self.days_spin)
        self.days_widget.setLayout(days_layout)
        date_layout.addWidget(self.days_widget)
        self.toggle_dates()
        
        controls_grid.addLayout(date_layout, row, 1)
        row += 1
        
        # Output file
        self.output_path_label = QLabel(self.translator.tr('output_path'))
        self.output_path_label.setFont(self.current_font)
        controls_grid.addWidget(self.output_path_label, row, 0)
        output_layout = QHBoxLayout()
        self.file_input = QLineEdit()
        self.file_input.setPlaceholderText(self.translator.tr('output_path'))
        self.file_input.setToolTip(self.translator.tr('tooltip_output_path'))
        self.file_input.setFont(self.current_font)
        output_layout.addWidget(self.file_input)
        self.browse_btn = QPushButton(self.translator.tr('browse_btn'))
        self.browse_btn.setIcon(QIcon.fromTheme('folder-open'))
        self.browse_btn.setToolTip(self.translator.tr('tooltip_browse'))
        self.browse_btn.setFont(self.current_font)
        self.browse_btn.clicked.connect(self.browse_file)
        output_layout.addWidget(self.browse_btn)
        controls_grid.addLayout(output_layout, row, 1)
        row += 1
        
        # Columns
        self.columns_label = QLabel(self.translator.tr('columns_label'))
        self.columns_label.setFont(self.current_font)
        controls_grid.addWidget(self.columns_label, row, 0)
        self.column_list = QListWidget()
        self.column_list.setToolTip(self.translator.tr('tooltip_columns'))
        self.column_list.setFont(self.current_font)
        for col in ["Date", "Open", "High", "Low", "Close", "Volume", "Spread", "RealVolume"]:
            item = QListWidgetItem(col)
            item.setCheckState(Qt.Checked)
            self.column_list.addItem(item)
        self.column_list.setMaximumHeight(150)
        controls_grid.addWidget(self.column_list, row, 1)
        row += 1
        
        # Plot selector
        self.plot_label = QLabel(self.translator.tr('plot_label'))
        self.plot_label.setFont(self.current_font)
        controls_grid.addWidget(self.plot_label, row, 0)
        self.symbol_combo = QComboBox()
        self.symbol_combo.setPlaceholderText(self.translator.tr('plot_label'))
        self.symbol_combo.setToolTip(self.translator.tr('tooltip_plot'))
        self.symbol_combo.setFont(self.current_font)
        controls_grid.addWidget(self.symbol_combo, row, 1)
        row += 1
        
        left_layout.addLayout(controls_grid)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFont(self.current_font)
        left_layout.addWidget(self.progress_bar)
        
        # Action buttons
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        
        self.download_btn = QPushButton(self.translator.tr('download_btn'))
        self.download_btn.setIcon(QIcon.fromTheme('document-download'))
        self.download_btn.setToolTip(self.translator.tr('tooltip_download'))
        self.download_btn.setFont(self.current_font)
        self.download_btn.clicked.connect(self.download_data)
        self.download_btn.setFixedHeight(48)
        button_layout.addWidget(self.download_btn)
        
        self.stop_btn = QPushButton(self.translator.tr('stop_btn'))
        self.stop_btn.setIcon(QIcon.fromTheme('process-stop'))
        self.stop_btn.setToolTip(self.translator.tr('tooltip_stop'))
        self.stop_btn.setFont(self.current_font)
        self.stop_btn.clicked.connect(self.stop_download)
        self.stop_btn.setEnabled(False)
        self.stop_btn.setFixedHeight(48)
        button_layout.addWidget(self.stop_btn)
        
        self.plot_btn = QPushButton(self.translator.tr('plot_btn'))
        self.plot_btn.setIcon(QIcon.fromTheme('office-chart-line'))
        self.plot_btn.setToolTip(self.translator.tr('tooltip_plot'))
        self.plot_btn.setFont(self.current_font)
        self.plot_btn.clicked.connect(self.plot_selected_symbol)
        self.plot_btn.setEnabled(False)
        self.plot_btn.setFixedHeight(48)
        button_layout.addWidget(self.plot_btn)
        
        left_layout.addLayout(button_layout)
        
        # Status messages
        self.status_label = QLabel()
        self.status_label.setWordWrap(True)
        status_font = QFont(self.current_font.family(), self.current_font.pointSize() - 2)
        self.status_label.setFont(status_font)
        left_layout.addWidget(self.status_label)
        
        left_panel.setLayout(left_layout)
        
        # Right panel (chart)
        right_panel = QWidget()
        right_layout = QVBoxLayout()
        right_layout.setContentsMargins(20, 20, 20, 20)
        
        self.chart = CandlestickChart(right_panel, width=8, height=6, dpi=100)
        self.toolbar = NavigationToolbar(self.chart, right_panel)
        right_layout.addWidget(self.toolbar)
        right_layout.addWidget(self.chart)
        
        right_panel.setLayout(right_layout)
        
        # Add panels to splitter
        main_splitter.addWidget(left_panel)
        main_splitter.addWidget(right_panel)
        main_splitter.setStretchFactor(0, 1)
        main_splitter.setStretchFactor(1, 2)
        
        main_layout.addWidget(main_splitter)
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)
        
        # Status bar
        self.statusBar().showMessage("Ready")

    def setup_menu_bar(self):
        """Set up the menu bar"""
        menu_bar = self.menuBar()
        
        # File menu
        file_menu = menu_bar.addMenu(self.translator.tr('menu_file'))
        save_preset_action = QAction(self.translator.tr('action_save_preset'), self)
        save_preset_action.setShortcut('Ctrl+S')
        save_preset_action.triggered.connect(self.save_symbol_preset)
        file_menu.addAction(save_preset_action)
        
        load_watchlist_action = QAction(self.translator.tr('action_load_watchlist'), self)
        load_watchlist_action.setShortcut('Ctrl+W')
        load_watchlist_action.triggered.connect(self.load_watchlist_symbols)
        file_menu.addAction(load_watchlist_action)
        
        file_menu.addSeparator()
        
        exit_action = QAction(self.translator.tr('action_exit'), self)
        exit_action.setShortcut('Ctrl+Q')
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # Settings menu
        settings_menu = menu_bar.addMenu(self.translator.tr('menu_settings'))
        
        select_font_action = QAction(self.translator.tr('action_select_font'), self)
        select_font_action.triggered.connect(self.select_font)
        settings_menu.addAction(select_font_action)
        
        font_size_menu = settings_menu.addMenu(self.translator.tr('action_font_size'))
        font_sizes = [8, 10, 12, 14, 16]
        self.font_size_actions = []
        for size in font_sizes:
            action = QAction(f"{size} pt", self, checkable=True)
            action.setData(size)
            action.triggered.connect(lambda checked, s=size: self.update_font(size=s))
            font_size_menu.addAction(action)
            self.font_size_actions.append(action)
            if size == self.current_font.pointSize():
                action.setChecked(True)
        
        settings_menu.addSeparator()
        
        chart_settings_action = QAction(self.translator.tr('action_chart_settings'), self)
        chart_settings_action.triggered.connect(self.open_chart_settings)
        settings_menu.addAction(chart_settings_action)
        
        toggle_language_action = QAction(self.translator.tr('action_toggle_language'), self)
        toggle_language_action.setShortcut('Ctrl+L')
        toggle_language_action.triggered.connect(self.toggle_language)
        settings_menu.addAction(toggle_language_action)
        
        # Help menu
        help_menu = menu_bar.addMenu(self.translator.tr('menu_help'))
        
        about_action = QAction(self.translator.tr('action_about'), self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)
        
        view_log_action = QAction(self.translator.tr('action_view_log'), self)
        view_log_action.setShortcut('Ctrl+Shift+L')
        view_log_action.triggered.connect(self.view_log)
        help_menu.addAction(view_log_action)

    def apply_dark_theme(self):
        """Apply flat dark theme to all UI elements"""
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor('#212121'))
        palette.setColor(QPalette.WindowText, QColor('#FFFFFF'))
        palette.setColor(QPalette.Base, QColor('#424242'))
        palette.setColor(QPalette.AlternateBase, QColor('#212121'))
        palette.setColor(QPalette.ToolTipBase, QColor('#424242'))
        palette.setColor(QPalette.ToolTipText, QColor('#FFFFFF'))
        palette.setColor(QPalette.Text, QColor('#FFFFFF'))
        palette.setColor(QPalette.Button, QColor('#6200EA'))
        palette.setColor(QPalette.ButtonText, QColor('#FFFFFF'))
        palette.setColor(QPalette.BrightText, QColor('#F44336'))
        palette.setColor(QPalette.Highlight, QColor('#6200EA'))
        palette.setColor(QPalette.HighlightedText, QColor('#FFFFFF'))
        
        # Menu bar styling
        menu_style = """
            QMenuBar {
                background-color: #212121;
                color: #FFFFFF;
                font-family: %s;
                font-size: %dpt;
            }
            QMenuBar::item {
                background-color: #212121;
                color: #FFFFFF;
                padding: 5px 10px;
            }
            QMenuBar::item:selected {
                background-color: #6200EA;
            }
            QMenu {
                background-color: #212121;
                color: #FFFFFF;
                border: 1px solid #616161;
                font-family: %s;
                font-size: %dpt;
            }
            QMenu::item {
                padding: 5px 20px;
            }
            QMenu::item:selected {
                background-color: #6200EA;
            }
        """ % (self.current_font.family(), self.current_font.pointSize(),
               self.current_font.family(), self.current_font.pointSize())
        self.menuBar().setStyleSheet(menu_style)
        
        # Button styling
        button_style = """
            QPushButton {
                background-color: #6200EA;
                color: #FFFFFF;
                border: none;
                padding: 10px;
                border-radius: 4px;
                font-family: %s;
                font-size: %dpt;
            }
            QPushButton:hover {
                background-color: #7C4DFF;
            }
            QPushButton:disabled {
                background-color: #616161;
                color: #B0BEC5;
            }
        """ % (self.current_font.family(), self.current_font.pointSize())
        self.download_btn.setStyleSheet(button_style)
        self.stop_btn.setStyleSheet(button_style)
        self.plot_btn.setStyleSheet(button_style)
        self.save_preset_btn.setStyleSheet(button_style)
        self.watchlist_btn.setStyleSheet(button_style)
        self.lang_btn.setStyleSheet(button_style)
        self.browse_btn.setStyleSheet(button_style)
        
        # Theme button
        theme_btn_style = """
            QToolButton {
                background-color: transparent;
                border: none;
                padding: 5px;
            }
            QToolButton:hover {
                background-color: #424242;
            }
        """
        self.theme_btn.setStyleSheet(theme_btn_style)
        
        # Toolbar styling
        toolbar_style = """
            QToolBar {
                background-color: #212121;
                border: none;
                spacing: 5px;
            }
            QToolButton {
                background-color: #424242;
                color: #FFFFFF;
                border: none;
                padding: 5px;
                border-radius: 4px;
                font-family: %s;
                font-size: %dpt;
            }
            QToolButton:hover {
                background-color: #6200EA;
            }
        """ % (self.current_font.family(), self.current_font.pointSize())
        self.toolbar.setStyleSheet(toolbar_style)
        
        # Input styling
        input_style = """
            QComboBox, QLineEdit, QSpinBox, QDateEdit {
                background-color: #424242;
                color: #FFFFFF;
                border: 1px solid #616161;
                padding: 8px;
                border-radius: 4px;
                font-family: %s;
                font-size: %dpt;
            }
            QComboBox QAbstractItemView {
                background-color: #424242;
                color: #FFFFFF;
                selection-background-color: #6200EA;
                border: none;
                font-family: %s;
                font-size: %dpt;
            }
        """ % (self.current_font.family(), self.current_font.pointSize(),
               self.current_font.family(), self.current_font.pointSize())
        self.symbol_input.setStyleSheet(input_style)
        self.format_combo.setStyleSheet(input_style)
        self.symbol_combo.setStyleSheet(input_style)
        self.days_spin.setStyleSheet(input_style)
        self.start_date.setStyleSheet(input_style)
        self.end_date.setStyleSheet(input_style)
        self.file_input.setStyleSheet(input_style)
        self.preset_combo.setStyleSheet(input_style)
        
        # List widget styling
        list_style = """
            QListWidget {
                background-color: #424242;
                color: #FFFFFF;
                border: 1px solid #616161;
                border-radius: 4px;
                font-family: %s;
                font-size: %dpt;
            }
            QListWidget::item {
                padding: 8px;
            }
            QListWidget::item:selected {
                background-color: #6200EA;
                color: #FFFFFF;
            }
            QListWidget::item:hover {
                background-color: #616161;
            }
        """ % (self.current_font.family(), self.current_font.pointSize())
        self.tf_list.setStyleSheet(list_style)
        self.column_list.setStyleSheet(list_style)
        
        # Labels
        label_style = """
            QLabel {
                color: #FFFFFF;
                font-family: %s;
                font-size: %dpt;
            }
        """ % (self.current_font.family(), self.current_font.pointSize())
        for widget in self.findChildren(QLabel):
            if widget != self.header_label and widget != self.status_label:
                widget.setStyleSheet(label_style)
        
        # Header label
        header_style = """
            QLabel {
                color: #FFFFFF;
                font-family: %s;
                font-size: %dpt;
                font-weight: bold;
            }
        """ % (self.current_font.family(), self.current_font.pointSize() + 4)
        self.header_label.setStyleSheet(header_style)
        
        # Status label
        status_style = """
            QLabel {
                color: #B0BEC5;
                font-family: %s;
                font-size: %dpt;
            }
        """ % (self.current_font.family(), self.current_font.pointSize() - 2)
        self.status_label.setStyleSheet(status_style)
        
        # Progress bar
        progress_style = """
            QProgressBar {
                background-color: #424242;
                color: #FFFFFF;
                border: none;
                border-radius: 4px;
                text-align: center;
                font-family: %s;
                font-size: %dpt;
                height: 24px;
            }
            QProgressBar::chunk {
                background-color: #6200EA;
                border-radius: 4px;
            }
        """ % (self.current_font.family(), self.current_font.pointSize())
        self.progress_bar.setStyleSheet(progress_style)
        
        QApplication.setPalette(palette)
        self.chart.dark_mode = self.dark_mode
        if hasattr(self, 'current_chart_symbol') and self.current_chart_symbol:
            self.plot_data(self.chart_data[self.current_chart_symbol], self.current_chart_symbol)

    def select_font(self):
        """Open font dialog to select a new font"""
        font, ok = QFontDialog.getFont(self.current_font, self, self.translator.tr('action_select_font'))
        if ok:
            self.update_font(font=font)
            self.statusBar().showMessage(f"Font changed to {font.family()}", 3000)

    def open_chart_settings(self):
        """Open chart settings dialog"""
        dialog = ChartSettingsDialog(self)
        if dialog.exec_():
            self.candle_colors['up'] = dialog.up_color.name()
            self.candle_colors['down'] = dialog.down_color.name()
            self.show_grid = dialog.grid_check.isChecked()
            self.show_volume = dialog.volume_check.isChecked()
            self.save_settings()
            if self.current_chart_symbol and self.current_chart_symbol in self.chart_data:
                self.plot_data(self.chart_data[self.current_chart_symbol], self.current_chart_symbol)
            self.statusBar().showMessage("Chart settings updated", 3000)

    def show_about(self):
        """Show about dialog"""
        QMessageBox.about(
            self,
            self.translator.tr('about_title'),
            self.translator.tr('about_text')
        )

    def view_log(self):
        """Open the log file in the default text editor"""
        log_file = 'mt5_downloader.log'
        if os.path.exists(log_file):
            QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.abspath(log_file)))
        else:
            QMessageBox.warning(self, "Error", "Log file not found.")

    def toggle_language(self):
        """Toggle between English and Farsi"""
        try:
            new_lang = 'fa' if self.translator.current_lang == 'en' else 'en'
            logging.info(f"Switching language to {new_lang}")
            if self.translator.set_language(new_lang):
                self.retranslate_ui()
                self.lang_btn.setText("FA" if new_lang == 'en' else "EN")
                self.statusBar().showMessage(f"Language switched to {new_lang.upper()}", 3000)
            else:
                logging.error(f"Failed to switch to language {new_lang}")
                QMessageBox.warning(self, "Error", f"Failed to switch to language {new_lang}")
        except Exception as e:
            logging.error(f"Error toggling language: {str(e)}")
            QMessageBox.warning(self, "Error", f"Failed to toggle language: {str(e)}")

    def retranslate_ui(self):
        """Update all UI elements with current language"""
        try:
            self.setWindowTitle(self.translator.tr('app_title'))
            
            # Update menu bar
            self.menuBar().clear()
            self.setup_menu_bar()
            
            # Update buttons
            self.download_btn.setText(self.translator.tr('download_btn'))
            self.stop_btn.setText(self.translator.tr('stop_btn'))
            self.plot_btn.setText(self.translator.tr('plot_btn'))
            self.save_preset_btn.setText(self.translator.tr('save_preset'))
            self.watchlist_btn.setText(self.translator.tr('load_watchlist'))
            self.browse_btn.setText(self.translator.tr('browse_btn'))
            
            # Update labels
            self.header_label.setText(self.translator.tr('app_title'))
            self.symbols_label.setText(self.translator.tr('symbols_label'))
            self.timeframes_label.setText(self.translator.tr('timeframes'))
            self.export_format_label.setText(self.translator.tr('export_format'))
            self.start_date_label.setText(self.translator.tr('start_date'))
            self.end_date_label.setText(self.translator.tr('end_date'))
            self.days_label.setText(self.translator.tr('days_label'))
            self.output_path_label.setText(self.translator.tr('output_path'))
            self.columns_label.setText(self.translator.tr('columns_label'))
            self.plot_label.setText(self.translator.tr('plot_label'))
            
            # Update radio buttons
            self.range_radio.setText(self.translator.tr('specific_range'))
            self.days_radio.setText(self.translator.tr('days_back'))
            
            # Update other elements
            self.preset_combo.setPlaceholderText(self.translator.tr('load_preset'))
            self.symbol_combo.setPlaceholderText(self.translator.tr('plot_label'))
            
            # Update tooltips
            self.symbol_input.setToolTip(self.translator.tr('tooltip_symbols'))
            self.tf_list.setToolTip(self.translator.tr('tooltip_timeframes'))
            self.format_combo.setToolTip(self.translator.tr('tooltip_export_format'))
            self.start_date.setToolTip(self.translator.tr('tooltip_date_range'))
            self.end_date.setToolTip(self.translator.tr('tooltip_date_range'))
            self.days_spin.setToolTip(self.translator.tr('tooltip_date_range'))
            self.file_input.setToolTip(self.translator.tr('tooltip_output_path'))
            self.column_list.setToolTip(self.translator.tr('tooltip_columns'))
            self.symbol_combo.setToolTip(self.translator.tr('tooltip_plot'))
            self.download_btn.setToolTip(self.translator.tr('tooltip_download'))
            self.stop_btn.setToolTip(self.translator.tr('tooltip_stop'))
            self.plot_btn.setToolTip(self.translator.tr('tooltip_plot'))
            self.save_preset_btn.setToolTip(self.translator.tr('tooltip_save_preset'))
            self.watchlist_btn.setToolTip(self.translator.tr('tooltip_load_watchlist'))
            self.browse_btn.setToolTip(self.translator.tr('tooltip_browse'))
            
            # RTL support for Farsi
            layout_direction = Qt.RightToLeft if self.translator.current_lang == 'fa' else Qt.LeftToRight
            self.setLayoutDirection(layout_direction)
            for widget in self.findChildren(QWidget):
                widget.setLayoutDirection(layout_direction)
            
            # Update chart
            if self.current_chart_symbol and self.current_chart_symbol in self.chart_data:
                self.plot_data(self.chart_data[self.current_chart_symbol], self.current_chart_symbol)
                
        except Exception as e:
            logging.error(f"Error in retranslate_ui: {str(e)}")
            raise

    def load_presets(self):
        """Load saved presets from config file"""
        self.preset_combo.clear()
        self.preset_combo.addItem(f"-- {self.translator.tr('load_preset')} --", None)
        try:
            if os.path.exists('symbol_presets.json'):
                with open('symbol_presets.json', 'r') as f:
                    presets = json.load(f)
                    for preset_name in presets.keys():
                        self.preset_combo.addItem(preset_name, presets[preset_name])
        except Exception as e:
            logging.error(f"Error loading presets: {e}")
            self.statusBar().showMessage("Error loading presets", 3000)

    def save_symbol_preset(self):
        """Save current symbols as a named preset"""
        current_symbols = self.symbol_input.text().strip()
        if not current_symbols:
            QMessageBox.warning(self, "Warning", self.translator.tr('no_symbols'))
            return
            
        preset_name, ok = QInputDialog.getText(
            self, self.translator.tr('save_preset'), self.translator.tr('preset_name')
        )
        if ok and preset_name:
            try:
                presets = {}
                if os.path.exists('symbol_presets.json'):
                    with open('symbol_presets.json', 'r') as f:
                        presets = json.load(f)
                
                presets[preset_name] = current_symbols
                
                with open('symbol_presets.json', 'w') as f:
                    json.dump(presets, f, indent=4)
                
                self.load_presets()
                self.statusBar().showMessage(
                    self.translator.tr('preset_saved').format(preset_name), 3000
                )
                
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to save preset: {e}")
                logging.error(f"Error saving preset: {e}")

    def on_preset_selected(self, index):
        """Load symbols from selected preset"""
        if index > 0:
            preset_symbols = self.preset_combo.currentData()
            if preset_symbols:
                self.symbol_input.setText(preset_symbols)
                self.statusBar().showMessage(
                    self.translator.tr('preset_loaded').format(self.preset_combo.currentText()), 
                    3000
                )

    def load_watchlist_symbols(self):
        """Load symbols from MT5 market watch"""
        if not mt5.initialize():
            QMessageBox.warning(self, "Error", self.translator.tr('mt5_connect_error'))
            return
        
        try:
            symbols = mt5.symbols_get()
            watchlist_symbols = [s.name for s in symbols if s.visible]
            
            if not watchlist_symbols:
                QMessageBox.information(self, "Info", self.translator.tr('no_watchlist'))
                return
                
            dialog = QDialog(self)
            dialog.setWindowTitle(self.translator.tr('watchlist_title'))
            dialog.setMinimumWidth(300)
            layout = QVBoxLayout()
            
            list_widget = QListWidget()
            list_widget.setSelectionMode(QListWidget.MultiSelection)
            list_widget.setFont(self.current_font)
            for symbol in sorted(watchlist_symbols):
                list_widget.addItem(symbol)
            
            for i in range(list_widget.count()):
                list_widget.item(i).setSelected(True)
            
            button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            button_box.accepted.connect(dialog.accept)
            button_box.rejected.connect(dialog.reject)
            
            layout.addWidget(QLabel(self.translator.tr('select_symbols')))
            layout.addWidget(list_widget)
            layout.addWidget(button_box)
            dialog.setLayout(layout)
            
            if dialog.exec_() == QDialog.Accepted:
                selected_items = [item.text() for item in list_widget.selectedItems()]
                if selected_items:
                    self.symbol_input.setText(",".join(sorted(selected_items)))
                    self.statusBar().showMessage(
                        self.translator.tr('watchlist_loaded').format(len(selected_items)), 
                        3000
                    )
                    
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to load watchlist: {e}")
            logging.error(f"Error loading watchlist: {e}")
        finally:
            mt5.shutdown()

    def get_current_symbols(self):
        """Get list of currently entered symbols"""
        return [s.strip() for s in self.symbol_input.text().split(',') if s.strip()]

    def validate_symbols(self, symbols):
        """Validate symbols against MT5's available symbols"""
        if not mt5.initialize():
            return False, self.translator.tr('mt5_connect_error')
        
        try:
            available_symbols = [s.name.lower() for s in mt5.symbols_get()]
            invalid_symbols = [
                s for s in symbols 
                if s.lower() not in available_symbols
            ]
            
            if invalid_symbols:
                return False, f"Invalid symbols: {', '.join(invalid_symbols)}"
            return True, "All symbols valid"
            
        except Exception as e:
            return False, str(e)
        finally:
            mt5.shutdown()

    def toggle_theme(self):
        """Toggle between dark and light themes (dark only as per request)"""
        self.theme_btn.setChecked(True)
        self.apply_dark_theme()
        QMessageBox.information(self, "Theme", self.translator.tr('theme_message'))

    def toggle_dates(self):
        self.date_widget.setVisible(self.range_radio.isChecked())
        self.days_widget.setVisible(self.days_radio.isChecked())

    def browse_file(self):
        fmt = self.format_combo.currentText()
        filter_str = "Excel Files (*.xlsx)" if fmt == "xlsx" else "CSV Files (*.csv)"
        file, _ = QFileDialog.getSaveFileName(self, self.translator.tr('browse_btn'), "", filter_str)
        if file:
            self.file_input.setText(file)

    def plot_selected_symbol(self):
        """Plot the currently selected symbol from combo box"""
        selected_symbol = self.symbol_combo.currentText()
        if not selected_symbol or selected_symbol not in self.chart_data:
            QMessageBox.warning(self, "No Data", "Please select a valid symbol to plot.")
            return
            
        data = self.chart_data[selected_symbol]
        self.current_chart_symbol = selected_symbol
        self.plot_data(data, selected_symbol)

    def plot_data(self, data, symbol=""):
        """Plot data for a specific symbol"""
        try:
            if data is None or data.empty:
                QMessageBox.warning(self, "No Data", self.translator.tr('no_data'))
                return
                
            plot_data = data.copy()
            
            if not pd.api.types.is_datetime64_any_dtype(plot_data['Date']):
                plot_data['Date'] = pd.to_datetime(plot_data['Date'])
            
            self.chart.plot_candles(
                plot_data, 
                symbol, 
                candle_colors=self.candle_colors,
                show_grid=self.show_grid,
                show_volume=self.show_volume
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Plot Error", f"Failed to plot data: {str(e)}")
            logging.error(f"Plot error: {str(e)}")

    def download_data(self):
        symbols = self.get_current_symbols()
        if not symbols:
            QMessageBox.warning(self, "Invalid Input", self.translator.tr('no_symbols'))
            return

        selected_timeframes = [item.text() for item in self.tf_list.selectedItems()]
        if not selected_timeframes:
            QMessageBox.warning(self, "Invalid Input", self.translator.tr('no_timeframes'))
            return

        if self.range_radio.isChecked():
            start = self.start_date.date().toPyDate()
            end = self.end_date.date().toPyDate()
            if start > end:
                QMessageBox.warning(self, "Invalid Range", self.translator.tr('invalid_range'))
                return
        else:
            end = datetime.now().date()
            start = end - timedelta(days=self.days_spin.value())
            end = min(end, datetime.now().date())

        export_format = self.format_combo.currentText()
        output_file = self.file_input.text().strip()
        
        selected_columns = [self.column_list.item(i).text() 
                          for i in range(self.column_list.count())
                          if self.column_list.item(i).checkState() == Qt.Checked]
        if not selected_columns:
            QMessageBox.warning(self, "Column Selection", self.translator.tr('no_columns'))
            return

        self.download_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.plot_btn.setEnabled(False)
        self.statusBar().showMessage(self.translator.tr('download_starting'))
        self.status_label.setText(self.translator.tr('download_starting'))
        
        self.download_thread = DataDownloadThread(
            symbols, selected_timeframes, 
            datetime.combine(start, datetime.min.time()),
            datetime.combine(end, datetime.min.time()),
            output_file, export_format, selected_columns
        )
        self.download_thread.progress.connect(self.progress_bar.setValue)
        self.download_thread.finished.connect(self.on_download_finished)
        self.download_thread.error.connect(self.download_error)
        self.download_thread.log_message.connect(self.update_status)
        self.download_thread.start()

    def stop_download(self):
        if hasattr(self, 'download_thread') and isinstance(self.download_thread, QThread) and self.download_thread.isRunning():
            self.download_thread.stop()
            self.status_label.setText("Download stopped by user")
            self.download_btn.setEnabled(True)
            self.stop_btn.setEnabled(False)
            self.statusBar().showMessage("Download stopped")

    def update_status(self, message, level):
        """Update status label with messages from the download thread"""
        color = {
            "INFO": "#B0BEC5",
            "WARNING": "#FFCA28",
            "ERROR": "#F44336"
        }.get(level, "#B0BEC5")
        
        self.status_label.setText(f"<font color='{color}'>{message}</font>")
        self.status_label.repaint()

    def on_download_finished(self, result_data):
        """Handle download completion"""
        self.chart_data = result_data
        self.download_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.progress_bar.setValue(100)
        
        self.symbol_combo.clear()
        if result_data:
            self.symbol_combo.addItems(sorted(result_data.keys()))
            self.plot_btn.setEnabled(True)
            self.statusBar().showMessage(self.translator.tr('download_complete'))
            self.status_label.setText(self.translator.tr('download_complete'))
            
            first_key = sorted(result_data.keys())[0]
            self.symbol_combo.setCurrentText(first_key)
            self.current_chart_symbol = first_key
            self.plot_data(result_data[first_key], first_key)
        else:
            self.statusBar().showMessage(self.translator.tr('no_data_received'))
            self.status_label.setText(self.translator.tr('no_data_received'))

    def download_error(self, error_msg):
        self.statusBar().showMessage(self.translator.tr('error_occurred'))
        self.status_label.setText(f"<font color='#F44336'>Error: {error_msg}</font>")
        self.download_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.plot_btn.setEnabled(False)
        self.progress_bar.setValue(0)

    def closeEvent(self, event):
        """Handle window close event"""
        if hasattr(self, 'download_thread') and isinstance(self.download_thread, QThread) and self.download_thread.isRunning():
            reply = QMessageBox.question(
                self, 'Download in Progress',
                self.translator.tr('confirm_exit'),
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                self.download_thread.stop()
                self.download_thread.wait(1000)
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()

def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    app.setStyleSheet("""
        QToolTip {
            background-color: #424242;
            color: #FFFFFF;
            border: none;
            padding: 5px;
            font-family: Roboto;
            font-size: 12pt;
        }
    """)
    
    window = MT5DataDownloader()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
